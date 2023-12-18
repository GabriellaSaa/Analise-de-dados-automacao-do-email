from openpyxl import load_workbook
from openpyxl.chart import BarChart
#print(BarChart)
from openpyxl.chart import Reference
#print(Reference)
import os
caminho_correto = os.path.join(
    "C:\\Users\\81035932",
    "OneDrive - Vale S.A",
    "Documents",
    "data",
    "pivot_table.xlsx")

#1- Lê pasta de trabalho e planilha
wb = load_workbook(caminho_correto)
#print(wb)
sheet = wb["Relatorio"]

#2 - Referências da s linhas e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row
#print(min_row)
#print(max_row)
barchart = BarChart()
#print(BarChart)
#Define as referências para a fonte, título e dados do gráfico
data = Reference(sheet,
    min_col = min_column + 1,
    max_col = max_column,
    min_row = min_row,  # Ajuste para pular a linha de categorias
    max_row = max_row,
)

categories = Reference(
    sheet,
    min_col = min_column,
    max_col = min_column,
    min_row = min_row + 1,
    max_row = min_row,
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

#4 - Criando o Gráfico

sheet.add_chart(barchart, "B10")
barchart.title = "Vendas por Fabricantes"
barchart.style = 2 

# 5 - Salvando o Worbook 
wb.save("barchart.xlsx")