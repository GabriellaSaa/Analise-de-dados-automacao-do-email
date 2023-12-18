from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

# 1- Lê pasta de trabalho e planilha
caminho_absoluto = os.path.join(
    "C:\\Users\\81035932",
    "OneDrive - Vale S.A",
    "Documents",
    "data",
    "barchart.xlsx"
)

wb = load_workbook(caminho_absoluto)
sheet = wb["Relatorio"]
print(wb)

#2 - Referências da s linhas e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#3 - Incluindo Fórmula
#sheet["B6"] = "=SUM(B2:B5)"
#sheet["B6"].style = "Currency"
for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    print(letter)
    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
    sheet[f"{letter}{max_row}"].style = "Currency"

wb.save("test.xlsx")